from __future__ import annotations

import argparse
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
import xlrd
from app.services.competition_source_resolver import (
    build_competition_source_manifest,
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()

    parser.add_argument(
        "--attachments",
        type=Path,
        required=True,
    )

    return parser.parse_args()


def main() -> None:
    args = parse_args()

    manifest = (
        build_competition_source_manifest(
            args.attachments
        )
    )

    excel_records = [
        record
        for record in manifest
        if record.source_type == "excel"
    ]

    extension_counts = Counter(
        record.extension
        for record in excel_records
    )

    parse_success = 0
    parse_failure = 0

    total_sheets = 0
    hidden_sheets = 0

    total_merged_ranges = 0
    workbooks_with_merged_cells = 0

    total_formula_cells = 0
    workbooks_with_formulas = 0

    max_rows_seen = 0
    max_columns_seen = 0

    sheet_count_distribution: Counter[
        int
    ] = Counter()

    failures: list[
        tuple[str, str]
    ] = []

    for index, record in enumerate(
        excel_records,
        start=1,
    ):
        path = (
            args.attachments
            / record.relative_path
        )

        print(
            f"[{index}/{len(excel_records)}] "
            f"{record.actual_filename}"
        )

        if record.extension == ".xls":
            try:
                workbook = xlrd.open_workbook(
                    path,
                    formatting_info=True,
                    on_demand=True,
                )

            except Exception as exc:
                parse_failure += 1

                failures.append(
                    (
                        record.actual_filename,
                        str(exc),
                    )
                )

                continue

            try:
                parse_success += 1

                sheet_count = workbook.nsheets

                total_sheets += sheet_count

                sheet_count_distribution[
                    sheet_count
                ] += 1

                workbook_has_merged = False

                for sheet_index in range(
                    workbook.nsheets
                ):
                    worksheet = (
                        workbook.sheet_by_index(
                            sheet_index
                        )
                    )

                    if worksheet.visibility != 0:
                        hidden_sheets += 1

                    max_rows_seen = max(
                        max_rows_seen,
                        worksheet.nrows,
                    )

                    max_columns_seen = max(
                        max_columns_seen,
                        worksheet.ncols,
                    )

                    merged_count = len(
                        worksheet.merged_cells
                    )

                    total_merged_ranges += (
                        merged_count
                    )

                    if merged_count > 0:
                        workbook_has_merged = True

                if workbook_has_merged:
                    workbooks_with_merged_cells += 1

            finally:
                workbook.release_resources()

            continue

        try:
            workbook = load_workbook(
                path,
                read_only=False,
                data_only=False,
            )

        except Exception as exc:
            parse_failure += 1

            failures.append(
                (
                    record.actual_filename,
                    str(exc),
                )
            )

            continue

        try:
            parse_success += 1

            sheet_count = len(
                workbook.worksheets
            )

            total_sheets += sheet_count

            sheet_count_distribution[
                sheet_count
            ] += 1

            workbook_has_merged = False
            workbook_has_formula = False

            for worksheet in (
                workbook.worksheets
            ):
                if (
                    worksheet.sheet_state
                    != "visible"
                ):
                    hidden_sheets += 1

                max_rows_seen = max(
                    max_rows_seen,
                    worksheet.max_row,
                )

                max_columns_seen = max(
                    max_columns_seen,
                    worksheet.max_column,
                )

                merged_count = len(
                    worksheet.merged_cells.ranges
                )

                total_merged_ranges += (
                    merged_count
                )

                if merged_count > 0:
                    workbook_has_merged = True

                sheet_formula_count = 0

                for row in (
                    worksheet.iter_rows()
                ):
                    for cell in row:
                        if (
                            cell.data_type == "f"
                        ):
                            sheet_formula_count += 1

                total_formula_cells += (
                    sheet_formula_count
                )

                if sheet_formula_count > 0:
                    workbook_has_formula = True

            if workbook_has_merged:
                workbooks_with_merged_cells += 1

            if workbook_has_formula:
                workbooks_with_formulas += 1

        finally:
            workbook.close()

    print()
    print(
        "=== Competition Excel Corpus Audit ==="
    )

    print(
        f"Excel files: {len(excel_records)}"
    )

    print(
        "Extensions:",
        dict(extension_counts),
    )

    print(
        f"Parse success: {parse_success}"
    )

    print(
        f"Parse failure: {parse_failure}"
    )

    print(
        f"Total sheets: {total_sheets}"
    )

    print(
        "Sheet count distribution:",
        dict(
            sorted(
                sheet_count_distribution.items()
            )
        ),
    )

    print(
        f"Hidden sheets: {hidden_sheets}"
    )

    print(
        "Workbooks with merged cells:",
        workbooks_with_merged_cells,
    )

    print(
        "Total merged ranges:",
        total_merged_ranges,
    )

    print(
        "XLSX workbooks with formulas:",
        workbooks_with_formulas,
    )

    print(
        "XLSX formula cells:",
        total_formula_cells,
    )

    print(
        "Note: XLS formula definitions "
        "are not counted by this audit."
    )

    print(
        f"Max rows seen: {max_rows_seen}"
    )

    print(
        f"Max columns seen: {max_columns_seen}"
    )

    if failures:
        print()
        print("Failures:")

        for filename, message in (
            failures[:30]
        ):
            print(
                f"- {filename}: "
                f"{message}"
            )

        if len(failures) > 30:
            print(
                f"... and "
                f"{len(failures) - 30} "
                f"more"
            )


if __name__ == "__main__":
    main()