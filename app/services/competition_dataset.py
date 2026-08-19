from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.schemas.competition import (
    CompetitionGold,
    CompetitionQaCase,
    CompetitionQuestion,
    CompetitionSolverInput,
    CompetitionSourceResolution,
)

_REQUIRED_COLUMNS = (
    "id",
    "source_type",
    "difficulty",
    "difficulty_cn",
    "qa_type",
    "question",
    "option_a",
    "option_b",
    "option_c",
    "option_d",
    "answer",
    "answer_text",
    "evidence",
    "source_title",
    "file_label",
)


class CompetitionDatasetError(
    RuntimeError
):
    pass


def _cell_to_text(
    value: object,
) -> str:
    if value is None:
        return ""

    if (
        isinstance(value, float)
        and value.is_integer()
    ):
        return str(int(value))

    return str(value).strip()


def load_competition_qa_excel(
    path: Path,
    *,
    worksheet_name: str | None = None,
) -> tuple[CompetitionQaCase, ...]:
    if not path.exists():
        raise CompetitionDatasetError(
            f"QA Excel 不存在: {path}"
        )

    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
    )

    try:
        if worksheet_name is None:
            worksheet = workbook.active
        else:
            if worksheet_name not in workbook.sheetnames:
                raise CompetitionDatasetError(
                    "QA Excel 中不存在工作表: "
                    f"{worksheet_name}"
                )

            worksheet = workbook[
                worksheet_name
            ]

        rows = worksheet.iter_rows(
            values_only=True,
        )

        try:
            raw_header = next(rows)
        except StopIteration as exc:
            raise CompetitionDatasetError(
                "QA Excel 为空"
            ) from exc

        header = tuple(
            _cell_to_text(value)
            for value in raw_header
        )

        column_index = {
            name: index
            for index, name
            in enumerate(header)
            if name
        }

        missing_columns = [
            column
            for column in _REQUIRED_COLUMNS
            if column not in column_index
        ]

        if missing_columns:
            raise CompetitionDatasetError(
                "QA Excel 缺少字段: "
                + ", ".join(
                    missing_columns
                )
            )

        cases: list[
            CompetitionQaCase
        ] = []

        seen_ids: set[str] = set()

        for row_number, row in enumerate(
            rows,
            start=2,
        ):
            case_id = _cell_to_text(
                row[
                    column_index["id"]
                ]
            )

            if not case_id:
                continue

            if case_id in seen_ids:
                raise CompetitionDatasetError(
                    "发现重复 QA ID: "
                    f"{case_id} "
                    f"(row={row_number})"
                )

            payload = {
                column: _cell_to_text(
                    row[
                        column_index[
                            column
                        ]
                    ]
                )
                for column
                in _REQUIRED_COLUMNS
            }

            payload["source_type"] = (
                payload["source_type"]
                .lower()
            )

            payload["difficulty"] = (
                payload["difficulty"]
                .lower()
            )

            payload["answer"] = (
                payload["answer"]
                .upper()
            )

            try:
                case = (
                    CompetitionQaCase(
                        case_id=payload.pop(
                            "id"
                        ),
                        **payload,
                    )
                )
            except Exception as exc:
                raise (
                    CompetitionDatasetError(
                        "QA 数据校验失败: "
                        f"row={row_number}, "
                        f"id={case_id}, "
                        f"error={exc}"
                    )
                ) from exc

            cases.append(case)
            seen_ids.add(case.case_id)

        if not cases:
            raise CompetitionDatasetError(
                "没有读取到任何 QA Case"
            )

        return tuple(cases)

    finally:
        workbook.close()

def build_competition_question(
    case: CompetitionQaCase,
) -> CompetitionQuestion:
    return CompetitionQuestion(
        case_id=case.case_id,
        source_type=case.source_type,
        qa_type=case.qa_type,
        question=case.question,
        option_a=case.option_a,
        option_b=case.option_b,
        option_c=case.option_c,
        option_d=case.option_d,
        source_title=case.source_title,
        file_label=case.file_label,
    )


def build_competition_gold(
    case: CompetitionQaCase,
) -> CompetitionGold:
    return CompetitionGold(
        case_id=case.case_id,
        source_type=case.source_type,
        qa_type=case.qa_type,
        difficulty=case.difficulty,
        difficulty_cn=(
            case.difficulty_cn
        ),
        answer=case.answer,
        answer_text=case.answer_text,
        evidence=case.evidence,
    )


def build_competition_solver_input(
    *,
    case: CompetitionQaCase,
    resolution: CompetitionSourceResolution,
) -> CompetitionSolverInput:
    if (
        case.case_id
        != resolution.case_id
    ):
        raise CompetitionDatasetError(
            "QA Case 与 Source Resolution "
            "case_id 不一致"
        )

    return CompetitionSolverInput(
        question=(
            build_competition_question(
                case
            )
        ),
        source=resolution,
    )