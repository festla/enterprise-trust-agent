from __future__ import annotations

from datetime import (
    date,
    datetime,
    time,
)
from decimal import Decimal
from pathlib import Path
from typing import Any

import xlrd
from openpyxl import (
    load_workbook,
)
from openpyxl.cell.cell import (
    MergedCell,
)
from openpyxl.utils import (
    get_column_letter,
)

from app.schemas.competition import (
    CompetitionSourceRecord,
)
from app.schemas.competition_excel import (
    CompetitionExcelCell,
    CompetitionExcelMergedRange,
    CompetitionExcelSheet,
    CompetitionExcelWorkbook,
)


class CompetitionExcelParserError(
    RuntimeError
):
    pass


def _coordinate(
    *,
    row_index: int,
    column_index: int,
) -> str:
    return (
        f"{get_column_letter(column_index)}"
        f"{row_index}"
    )


def _primitive_value(
    value: Any,
) -> (
    int
    | float
    | str
    | bool
    | None
):
    if value is None:
        return None

    if isinstance(value, bool):
        return value

    if isinstance(value, Decimal):
        return float(value)

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        return value

    if isinstance(
        value,
        (
            datetime,
            date,
            time,
        ),
    ):
        return value.isoformat()

    if isinstance(value, str):
        return value

    return str(value)


def _text_value(
    value: Any,
) -> str:
    if value is None:
        return ""

    if isinstance(value, bool):
        return (
            "TRUE"
            if value
            else "FALSE"
        )

    if isinstance(
        value,
        (
            datetime,
            date,
            time,
        ),
    ):
        return value.isoformat()

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))

    return str(value).strip()


def _numeric_value(
    value: Any,
) -> int | float | None:
    # bool 是 int 的子类，
    # 所以必须先排除。
    if isinstance(value, bool):
        return None

    if isinstance(value, Decimal):
        return float(value)

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        return value

    return None


# ============================================================
# XLSX / openpyxl
# ============================================================


def _xlsx_merged_ranges(
    worksheet,
) -> tuple[
    tuple[
        CompetitionExcelMergedRange,
        ...
    ],
    dict[str, str],
]:
    ranges: list[
        CompetitionExcelMergedRange
    ] = []

    anchor_to_range: dict[
        str,
        str,
    ] = {}

    for merged in (
        worksheet.merged_cells.ranges
    ):
        min_row = merged.min_row
        max_row = merged.max_row

        min_column = merged.min_col
        max_column = merged.max_col

        anchor = _coordinate(
            row_index=min_row,
            column_index=min_column,
        )

        range_ref = str(merged)

        ranges.append(
            CompetitionExcelMergedRange(
                range_ref=range_ref,
                min_row=min_row,
                max_row=max_row,
                min_column=min_column,
                max_column=max_column,
                anchor_coordinate=anchor,
            )
        )

        anchor_to_range[
            anchor
        ] = range_ref

    return (
        tuple(ranges),
        anchor_to_range,
    )


def _xlsx_cell_type(
    cell,
) -> str:
    if cell.data_type == "f":
        return "formula"

    if cell.data_type == "b":
        return "boolean"

    if cell.data_type == "e":
        return "error"

    if cell.is_date:
        return "date"

    if isinstance(
        cell.value,
        (int, float, Decimal),
    ) and not isinstance(
        cell.value,
        bool,
    ):
        return "number"

    return "text"


def _build_xlsx_cell(
    *,
    cell,
    cached_cell,
    merged_range: str | None,
) -> CompetitionExcelCell | None:
    if isinstance(
        cell,
        MergedCell,
    ):
        return None

    value = cell.value

    if value is None:
        return None

    if (
        isinstance(value, str)
        and not value.strip()
    ):
        return None

    coordinate = cell.coordinate

    if cell.data_type == "f":
        cached_value = (
            cached_cell.value
            if cached_cell is not None
            else None
        )

        effective_value = (
            cached_value
            if cached_value is not None
            else value
        )

        return CompetitionExcelCell(
            coordinate=coordinate,
            row_index=cell.row,
            column_index=cell.column,
            cell_type="formula",
            text_value=_text_value(
                effective_value
            ),
            numeric_value=(
                _numeric_value(
                    cached_value
                )
            ),
            number_format=(
                cell.number_format
            ),
            formula=str(value),
            formula_cached_value=(
                _primitive_value(
                    cached_value
                )
            ),
            merged_range=merged_range,
        )

    return CompetitionExcelCell(
        coordinate=coordinate,
        row_index=cell.row,
        column_index=cell.column,
        cell_type=_xlsx_cell_type(
            cell
        ),
        text_value=_text_value(
            value
        ),
        numeric_value=(
            _numeric_value(value)
        ),
        number_format=(
            cell.number_format
        ),
        merged_range=merged_range,
    )


def _parse_xlsx(
    *,
    path: Path,
    source: CompetitionSourceRecord,
) -> CompetitionExcelWorkbook:
    # 第一份 Workbook：
    # 保存公式定义。
    formula_workbook = load_workbook(
        path,
        read_only=False,
        data_only=False,
    )

    # 第二份 Workbook：
    # 用来读取公式缓存值。
    value_workbook = load_workbook(
        path,
        read_only=False,
        data_only=True,
    )

    try:
        sheets: list[
            CompetitionExcelSheet
        ] = []

        for (
            sheet_index,
            worksheet,
        ) in enumerate(
            formula_workbook.worksheets
        ):
            cached_worksheet = (
                value_workbook[
                    worksheet.title
                ]
            )

            (
                merged_ranges,
                anchor_to_range,
            ) = _xlsx_merged_ranges(
                worksheet
            )

            cells: list[
                CompetitionExcelCell
            ] = []

            # ================================================
            # 这里故意使用 openpyxl 的 Sparse Cell Registry。
            #
            # 不允许：
            #
            # for row in range(ws.max_row):
            #     for col in range(ws.max_column):
            #
            # 因为真实数据已经出现：
            #
            # 10007 x 16384
            #
            # _cells 是 openpyxl 内部结构，
            # 但当前比赛 MVP 中它能避免对巨大声明范围
            # 进行无意义扫描。
            # ================================================

            sparse_cells = list(
                worksheet._cells.values()
            )

            for cell in sparse_cells:
                if isinstance(
                    cell,
                    MergedCell,
                ):
                    continue

                cached_cell = None

                if (
                    cell.data_type
                    == "f"
                ):
                    cached_cell = (
                        cached_worksheet[
                            cell.coordinate
                        ]
                    )

                parsed = (
                    _build_xlsx_cell(
                        cell=cell,
                        cached_cell=(
                            cached_cell
                        ),
                        merged_range=(
                            anchor_to_range.get(
                                cell.coordinate
                            )
                        ),
                    )
                )

                if parsed is not None:
                    cells.append(parsed)

            cells.sort(
                key=lambda item: (
                    item.row_index,
                    item.column_index,
                )
            )

            sheets.append(
                CompetitionExcelSheet(
                    name=worksheet.title,
                    sheet_index=sheet_index,
                    visible=(
                        worksheet.sheet_state
                        == "visible"
                    ),
                    declared_max_row=(
                        worksheet.max_row
                    ),
                    declared_max_column=(
                        worksheet.max_column
                    ),
                    cells=tuple(cells),
                    merged_ranges=(
                        merged_ranges
                    ),
                )
            )

        return CompetitionExcelWorkbook(
            source_id=source.source_id,
            relative_path=(
                source.relative_path
            ),
            excel_format="xlsx",
            sheets=tuple(sheets),
        )

    finally:
        formula_workbook.close()
        value_workbook.close()


# ============================================================
# XLS / xlrd
# ============================================================


def _xls_number_format(
    *,
    workbook,
    cell,
) -> str | None:
    try:
        xf = workbook.xf_list[
            cell.xf_index
        ]

        format_info = (
            workbook.format_map.get(
                xf.format_key
            )
        )

        if format_info is None:
            return None

        return format_info.format_str

    except Exception:
        # number_format 只是增强字段。
        # 读取失败不能导致整个 Cell Parser 失败。
        return None


def _xls_merged_ranges(
    worksheet,
) -> tuple[
    tuple[
        CompetitionExcelMergedRange,
        ...
    ],
    dict[str, str],
]:
    ranges: list[
        CompetitionExcelMergedRange
    ] = []

    anchor_to_range: dict[
        str,
        str,
    ] = {}

    # xlrd:
    #
    # (row_low, row_high,
    #  col_low, col_high)
    #
    # high 是 exclusive。
    for (
        row_low,
        row_high,
        col_low,
        col_high,
    ) in worksheet.merged_cells:
        min_row = row_low + 1
        max_row = row_high

        min_column = col_low + 1
        max_column = col_high

        anchor = _coordinate(
            row_index=min_row,
            column_index=min_column,
        )

        end = _coordinate(
            row_index=max_row,
            column_index=max_column,
        )

        range_ref = (
            f"{anchor}:{end}"
        )

        ranges.append(
            CompetitionExcelMergedRange(
                range_ref=range_ref,
                min_row=min_row,
                max_row=max_row,
                min_column=min_column,
                max_column=max_column,
                anchor_coordinate=anchor,
            )
        )

        anchor_to_range[
            anchor
        ] = range_ref

    return (
        tuple(ranges),
        anchor_to_range,
    )


def _build_xls_cell(
    *,
    workbook,
    cell,
    row_index: int,
    column_index: int,
    merged_range: str | None,
) -> CompetitionExcelCell | None:
    cell_type = cell.ctype

    if cell_type in {
        xlrd.XL_CELL_EMPTY,
        xlrd.XL_CELL_BLANK,
    }:
        return None

    coordinate = _coordinate(
        row_index=row_index,
        column_index=column_index,
    )

    value = cell.value

    if (
        isinstance(value, str)
        and not value.strip()
    ):
        return None

    if cell_type == xlrd.XL_CELL_TEXT:
        parsed_type = "text"

        text = _text_value(value)

        numeric = None

    elif cell_type == xlrd.XL_CELL_NUMBER:
        parsed_type = "number"

        text = _text_value(value)

        numeric = _numeric_value(
            value
        )

    elif cell_type == xlrd.XL_CELL_DATE:
        parsed_type = "date"

        try:
            converted = (
                xlrd.xldate_as_datetime(
                    value,
                    workbook.datemode,
                )
            )

            text = (
                converted.isoformat()
            )

        except Exception:
            text = _text_value(
                value
            )

        numeric = None

    elif (
        cell_type
        == xlrd.XL_CELL_BOOLEAN
    ):
        parsed_type = "boolean"

        boolean_value = bool(value)

        text = (
            "TRUE"
            if boolean_value
            else "FALSE"
        )

        numeric = None

    elif (
        cell_type
        == xlrd.XL_CELL_ERROR
    ):
        parsed_type = "error"

        text = (
            xlrd.error_text_from_code.get(
                value,
                str(value),
            )
        )

        numeric = None

    else:
        parsed_type = "text"

        text = _text_value(value)

        numeric = None

    return CompetitionExcelCell(
        coordinate=coordinate,
        row_index=row_index,
        column_index=column_index,
        cell_type=parsed_type,
        text_value=text,
        numeric_value=numeric,
        number_format=(
            _xls_number_format(
                workbook=workbook,
                cell=cell,
            )
        ),
        merged_range=merged_range,
    )


def _parse_xls(
    *,
    path: Path,
    source: CompetitionSourceRecord,
) -> CompetitionExcelWorkbook:
    workbook = xlrd.open_workbook(
        path,
        formatting_info=True,
        on_demand=True,
    )

    try:
        sheets: list[
            CompetitionExcelSheet
        ] = []

        for sheet_index in range(
            workbook.nsheets
        ):
            worksheet = (
                workbook.sheet_by_index(
                    sheet_index
                )
            )

            (
                merged_ranges,
                anchor_to_range,
            ) = _xls_merged_ranges(
                worksheet
            )

            cells: list[
                CompetitionExcelCell
            ] = []

            # XLS 最大列数远小于 XLSX。
            # row() 同时避免我们根据某个夸张的
            # 16384 列声明范围构造完整矩阵。
            for row_zero_index in range(
                worksheet.nrows
            ):
                row_cells = worksheet.row(
                    row_zero_index
                )

                for (
                    col_zero_index,
                    cell,
                ) in enumerate(
                    row_cells
                ):
                    row_index = (
                        row_zero_index + 1
                    )

                    column_index = (
                        col_zero_index + 1
                    )

                    coordinate = (
                        _coordinate(
                            row_index=(
                                row_index
                            ),
                            column_index=(
                                column_index
                            ),
                        )
                    )

                    parsed = (
                        _build_xls_cell(
                            workbook=workbook,
                            cell=cell,
                            row_index=(
                                row_index
                            ),
                            column_index=(
                                column_index
                            ),
                            merged_range=(
                                anchor_to_range.get(
                                    coordinate
                                )
                            ),
                        )
                    )

                    if parsed is not None:
                        cells.append(
                            parsed
                        )

            sheets.append(
                CompetitionExcelSheet(
                    name=worksheet.name,
                    sheet_index=sheet_index,
                    visible=(
                        worksheet.visibility
                        == 0
                    ),
                    declared_max_row=(
                        worksheet.nrows
                    ),
                    declared_max_column=(
                        worksheet.ncols
                    ),
                    cells=tuple(cells),
                    merged_ranges=(
                        merged_ranges
                    ),
                )
            )

        return CompetitionExcelWorkbook(
            source_id=source.source_id,
            relative_path=(
                source.relative_path
            ),
            excel_format="xls",
            sheets=tuple(sheets),
        )

    finally:
        workbook.release_resources()


# ============================================================
# Public API
# ============================================================


def parse_competition_excel(
    *,
    attachments_root: Path,
    source: CompetitionSourceRecord,
) -> CompetitionExcelWorkbook:
    if source.source_type != "excel":
        raise CompetitionExcelParserError(
            "CompetitionExcelParser "
            "只能处理 excel source"
        )

    path = (
        attachments_root
        / source.relative_path
    )

    if not path.exists():
        raise CompetitionExcelParserError(
            f"Excel 文件不存在: {path}"
        )

    if source.extension == ".xlsx":
        try:
            return _parse_xlsx(
                path=path,
                source=source,
            )
        except Exception as exc:
            raise CompetitionExcelParserError(
                "XLSX 解析失败: "
                f"{source.relative_path}: "
                f"{exc}"
            ) from exc

    if source.extension == ".xls":
        try:
            return _parse_xls(
                path=path,
                source=source,
            )
        except Exception as exc:
            raise CompetitionExcelParserError(
                "XLS 解析失败: "
                f"{source.relative_path}: "
                f"{exc}"
            ) from exc

    raise CompetitionExcelParserError(
        "不支持的 Excel 格式: "
        f"{source.extension}"
    )